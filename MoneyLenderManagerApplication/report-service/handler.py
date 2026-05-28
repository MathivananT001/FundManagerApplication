"""Report Service Lambda handlers — generates PDF/Excel reports, uploads to S3 (US-028, US-029, US-030)."""
import json
import os
import io
from datetime import datetime

import boto3
import httpx
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from openpyxl import Workbook

s3 = boto3.client("s3", region_name=os.environ.get("AWS_REGION", "ap-south-1"))
BUCKET = os.environ.get("S3_REPORTS_BUCKET", "mlm-reports-dev")
PRESIGNED_EXPIRY = int(os.environ.get("PRESIGNED_URL_EXPIRY", "900"))
GROUP_SERVICE_URL = os.environ.get("GROUP_SERVICE_URL", "http://group-service:8000")
PAYMENT_SERVICE_URL = os.environ.get("PAYMENT_SERVICE_URL", "http://payment-service:8000")
AUCTION_SERVICE_URL = os.environ.get("AUCTION_SERVICE_URL", "http://auction-service:8000")


def generate_group_summary_report(event, context):
    """Generate group summary report (US-028, US-030)."""
    body = json.loads(event.get("body", "{}")) if isinstance(event.get("body"), str) else event
    group_id = body["group_id"]
    fmt = body.get("format", "pdf")

    # Fetch group data
    group = _fetch_json(f"{GROUP_SERVICE_URL}/groups/{group_id}")
    if not group:
        return _error_response("Group not found")

    if fmt == "pdf":
        content = _build_group_summary_pdf(group)
        content_type = "application/pdf"
        ext = "pdf"
    else:
        content = _build_group_summary_excel(group)
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ext = "xlsx"

    s3_key = f"reports/{group_id}/group-summary-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.{ext}"
    s3.put_object(Bucket=BUCKET, Key=s3_key, Body=content, ContentType=content_type)
    url = _create_presigned_url(s3_key)

    return {"statusCode": 200, "body": json.dumps({"download_url": url, "s3_key": s3_key})}


def generate_member_history_report(event, context):
    """Generate member contribution history report (US-029, US-030)."""
    body = json.loads(event.get("body", "{}")) if isinstance(event.get("body"), str) else event
    group_id = body["group_id"]
    member_id = body["member_id"]
    fmt = body.get("format", "pdf")

    # Fetch data from payment service (all months)
    payments = []
    for month in range(1, 16):
        ledger = _fetch_json(f"{PAYMENT_SERVICE_URL}/payments/{group_id}/{month}")
        if ledger:
            for entry in ledger:
                if entry.get("member_id") == member_id:
                    payments.append(entry)

    if fmt == "pdf":
        content = _build_member_history_pdf(member_id, payments)
        content_type = "application/pdf"
        ext = "pdf"
    else:
        content = _build_member_history_excel(member_id, payments)
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ext = "xlsx"

    s3_key = f"reports/{group_id}/{member_id}/history-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.{ext}"
    s3.put_object(Bucket=BUCKET, Key=s3_key, Body=content, ContentType=content_type)
    url = _create_presigned_url(s3_key)

    return {"statusCode": 200, "body": json.dumps({"download_url": url, "s3_key": s3_key})}


def generate_auction_history_report(event, context):
    """Generate auction history report (US-030)."""
    body = json.loads(event.get("body", "{}")) if isinstance(event.get("body"), str) else event
    group_id = body["group_id"]
    fmt = body.get("format", "pdf")

    auctions = _fetch_json(f"{AUCTION_SERVICE_URL}/auctions/group/{group_id}") or []

    if fmt == "pdf":
        content = _build_auction_history_pdf(group_id, auctions)
        content_type = "application/pdf"
        ext = "pdf"
    else:
        content = _build_auction_history_excel(group_id, auctions)
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ext = "xlsx"

    s3_key = f"reports/{group_id}/auction-history-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.{ext}"
    s3.put_object(Bucket=BUCKET, Key=s3_key, Body=content, ContentType=content_type)
    url = _create_presigned_url(s3_key)

    return {"statusCode": 200, "body": json.dumps({"download_url": url, "s3_key": s3_key})}


# --- PDF Builders ---

def _build_group_summary_pdf(group: dict) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f"Group Summary: {group.get('name', '')}", styles["Title"]),
        Spacer(1, 12),
        Paragraph(f"Members: {group.get('member_count', 0)} / {group.get('member_slots', 0)}", styles["Normal"]),
        Paragraph(f"Targeting Amount: ₹{group.get('targeting_amount', 0):,.2f}", styles["Normal"]),
        Paragraph(f"Monthly Auction Amount: ₹{group.get('monthly_auction_amount', 0):,.2f}", styles["Normal"]),
        Paragraph(f"Manager Fee: {group.get('manager_fee_percent', 0)}%", styles["Normal"]),
        Paragraph(f"Status: {group.get('status', '')}", styles["Normal"]),
    ]
    doc.build(elements)
    return buf.getvalue()


def _build_member_history_pdf(member_id: str, payments: list) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f"Member Contribution History: {member_id}", styles["Title"]),
        Spacer(1, 12),
    ]
    if payments:
        data = [["Month", "Amount Due", "Status"]]
        for p in payments:
            data.append([str(p.get("month_number")), f"₹{p.get('amount_due', 0):,.2f}", p.get("payment_status", "")])
        t = Table(data)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(t)
    else:
        elements.append(Paragraph("No payment records found.", styles["Normal"]))
    doc.build(elements)
    return buf.getvalue()


def _build_auction_history_pdf(group_id: str, auctions: list) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f"Auction History: {group_id}", styles["Title"]),
        Spacer(1, 12),
    ]
    if auctions:
        data = [["Month", "Status", "Winner", "Winning Bid", "Disbursement"]]
        for a in auctions:
            data.append([
                str(a.get("month_number")), a.get("status", ""),
                a.get("winner_id", "—"),
                f"₹{a.get('winning_bid_amount', 0) or 0:,.2f}",
                f"₹{a.get('disbursement_amount', 0) or 0:,.2f}",
            ])
        t = Table(data)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(t)
    doc.build(elements)
    return buf.getvalue()


# --- Excel Builders ---

def _build_group_summary_excel(group: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Group Summary"
    ws.append(["Field", "Value"])
    ws.append(["Name", group.get("name", "")])
    ws.append(["Members", f"{group.get('member_count', 0)} / {group.get('member_slots', 0)}"])
    ws.append(["Targeting Amount (₹)", group.get("targeting_amount", 0)])
    ws.append(["Monthly Auction Amount (₹)", group.get("monthly_auction_amount", 0)])
    ws.append(["Manager Fee (%)", group.get("manager_fee_percent", 0)])
    ws.append(["Status", group.get("status", "")])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_member_history_excel(member_id: str, payments: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Member History"
    ws.append(["Month", "Amount Due (₹)", "Status"])
    for p in payments:
        ws.append([p.get("month_number"), p.get("amount_due", 0), p.get("payment_status", "")])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_auction_history_excel(group_id: str, auctions: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Auction History"
    ws.append(["Month", "Status", "Winner", "Winning Bid (₹)", "Disbursement (₹)"])
    for a in auctions:
        ws.append([
            a.get("month_number"), a.get("status"),
            a.get("winner_id", ""), a.get("winning_bid_amount", 0), a.get("disbursement_amount", 0),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- Helpers ---

def _fetch_json(url: str) -> dict | list | None:
    try:
        resp = httpx.get(url, timeout=10)
        return resp.json() if resp.status_code == 200 else None
    except Exception:
        return None


def _create_presigned_url(s3_key: str) -> str:
    return s3.generate_presigned_url(
        "get_object", Params={"Bucket": BUCKET, "Key": s3_key}, ExpiresIn=PRESIGNED_EXPIRY,
    )


def _error_response(message: str) -> dict:
    return {"statusCode": 400, "body": json.dumps({"error": message})}
